"""
Reporte de Tareas · IMEMSA

App de una sola pantalla: lee Google Tasks de la cuenta de CADA
gerente (una cuenta por persona, cada una autorizada individualmente
con su propio refresh_token) y genera un Excel consolidado.

Es de SOLO LECTURA — nunca escribe ni modifica nada en Tasks.

Requiere en .streamlit/secrets.toml (generado automáticamente por
armar_secrets.py, ver ese script):

[google_oauth]
client_id = "..."
client_secret = "..."

[[gerentes]]
nombre = "Rodrigo Herrera"
refresh_token = "..."

[[gerentes]]
nombre = "Florentino Pérez"
refresh_token = "..."
"""

import io
import time
from datetime import datetime, date

import streamlit as st
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build

NAVY = "0D2B6E"
RED = "C41E2E"

st.set_page_config(page_title="IMEMSA · Reporte de Tareas", page_icon="📊", layout="centered")

st.markdown("""
<style>
    .stApp { background-color: #f7f8fb; }
    div.stButton > button:first-child, div.stDownloadButton > button:first-child {
        background-color: #0D2B6E; color: white; border-radius: 6px; border: none;
        font-weight: 600; padding: 0.55em 1.4em;
    }
    div.stButton > button:first-child:hover, div.stDownloadButton > button:first-child:hover {
        background-color: #C41E2E; color: white;
    }
    .imemsa-header {
        background: linear-gradient(135deg, #0D2B6E 0%, #123a8f 100%);
        border-radius: 12px;
        padding: 28px 32px;
        margin-bottom: 28px;
        box-shadow: 0 4px 14px rgba(13, 43, 110, 0.18);
    }
    .imemsa-header .eyebrow {
        color: #9db3e8; font-size: 0.78em; font-weight: 700;
        letter-spacing: 2px; text-transform: uppercase; margin-bottom: 6px;
    }
    .imemsa-header h1 {
        color: white; font-size: 1.9em; font-weight: 800; margin: 0 0 6px 0;
    }
    .imemsa-header p {
        color: #cfdaf5; font-size: 0.98em; margin: 0;
    }
</style>

<div class="imemsa-header">
    <div class="eyebrow">IMEMSA · Planta</div>
    <h1>📊 Reporte de Tareas por Gerente</h1>
    <p>Consolida en un clic las tareas capturadas por cada gerente en Google Tasks,
    con fechas, estatus y semáforo — listo para exportar a Excel.</p>
</div>
""", unsafe_allow_html=True)

_mostrar_debug = st.query_params.get("debug") == "1"


# --------------------------------------------------------------
# Conexión de solo lectura a Google Tasks, una por gerente
# --------------------------------------------------------------
def obtener_servicio(refresh_token: str):
    """Se reconstruye en cada llamada (no se cachea) para evitar reutilizar
    una conexión de red que haya quedado rota (ej. error 'Broken pipe')."""
    cfg = st.secrets["google_oauth"]
    creds = Credentials(
        token=None,
        refresh_token=refresh_token,
        client_id=cfg["client_id"],
        client_secret=cfg["client_secret"],
        token_uri="https://oauth2.googleapis.com/token",
        scopes=["https://www.googleapis.com/auth/tasks.readonly"],
    )
    return build("tasks", "v1", credentials=creds, cache_discovery=False)


def mostrar_debug_credenciales():
    """Muestra, de forma segura (sin exponer los valores completos), qué
    credenciales está leyendo la app AHORA MISMO — para comparar contra
    Google Cloud Console / gerentes_tokens.json sin adivinar."""
    cfg_oauth = st.secrets.get("google_oauth", {})
    gerentes = st.secrets.get("gerentes", [])
    with st.expander("🔧 Ver qué credenciales está usando la app (debug)"):
        st.code(f"client_id completo: {cfg_oauth.get('client_id', '(no encontrado)')}", language="text")
        st.write(f"**Gerentes configurados: {len(gerentes)}**")
        for g in gerentes:
            rt = g.get("refresh_token", "")
            st.code(f"{g.get('nombre', '(sin nombre)')}: {rt[:12]}... ({len(rt)} caracteres)", language="text")


def parsear_fecha(valor_rfc3339):
    if not valor_rfc3339:
        return None
    return datetime.fromisoformat(valor_rfc3339.replace("Z", "+00:00")).date()


def calcular_semaforo(estado, fecha_vencimiento, fecha_completado, hoy=None):
    hoy = hoy or date.today()
    if estado == "completed":
        if fecha_completado and fecha_vencimiento:
            return "Cerrada en tiempo" if fecha_completado <= fecha_vencimiento else "Cerrada con atraso"
        return "Cerrada"
    if fecha_vencimiento:
        dias = (fecha_vencimiento - hoy).days
        if dias < 0:
            return f"Vencida ({-dias} días de atraso)"
        elif dias <= 3:
            return f"Por vencer ({dias} días)"
        return "En tiempo"
    return "Sin fecha"


def _extraer_tareas_de_cuenta(nombre_gerente: str, refresh_token: str):
    """Extrae todas las tareas de UNA cuenta (todas sus listas)."""
    ultimo_error = None
    for intento in range(1, 4):
        try:
            servicio = obtener_servicio(refresh_token)
            listas = servicio.tasklists().list(maxResults=100).execute().get("items", [])

            filas = []
            for lista in listas:
                pagina = None
                while True:
                    resp = servicio.tasks().list(
                        tasklist=lista["id"], showCompleted=True, showHidden=True, showDeleted=False,
                        maxResults=100, pageToken=pagina,
                    ).execute()
                    for t in resp.get("items", []):
                        fecha_venc = parsear_fecha(t.get("due"))
                        fecha_completado = parsear_fecha(t.get("completed"))
                        filas.append({
                            "Gerente": nombre_gerente,
                            "Lista": lista["title"],
                            "Tarea": t.get("title", "(sin título)"),
                            "Notas": t.get("notes", ""),
                            "Fecha de vencimiento": fecha_venc,
                            "Estado Google Tasks": "Completada" if t.get("status") == "completed" else "Pendiente",
                            "Fecha de completado": fecha_completado,
                            "Semáforo": calcular_semaforo(t.get("status"), fecha_venc, fecha_completado),
                        })
                    pagina = resp.get("nextPageToken")
                    if not pagina:
                        break
            return filas, None
        except (BrokenPipeError, ConnectionError, OSError) as e:
            ultimo_error = e
            time.sleep(1.5 * intento)
        except Exception as e:
            return [], str(e)
    return [], str(ultimo_error)


@st.cache_data(ttl=120, show_spinner=False)
def extraer_todas_las_cuentas():
    """Recorre TODOS los gerentes configurados y consolida sus tareas.
    Si una cuenta falla, se reporta el error para esa cuenta pero se
    sigue con las demás — un problema no tumba el reporte completo."""
    gerentes = st.secrets.get("gerentes", [])
    todas_las_filas = []
    errores = []

    for g in gerentes:
        filas, error = _extraer_tareas_de_cuenta(g["nombre"], g["refresh_token"])
        if error:
            errores.append(f"{g['nombre']}: {error}")
        todas_las_filas.extend(filas)

    return pd.DataFrame(todas_las_filas), errores


# --------------------------------------------------------------
# Métricas ejecutivas para el Dashboard
# --------------------------------------------------------------
def calcular_metricas(df: pd.DataFrame) -> dict:
    hoy = date.today()

    cerradas = df[df["Estado Google Tasks"] == "Completada"]
    abiertas = df[df["Estado Google Tasks"] == "Pendiente"]

    cerradas_en_tiempo = cerradas[cerradas["Semáforo"] == "Cerrada en tiempo"]
    pct_efectividad_global = (
        round(100 * len(cerradas_en_tiempo) / len(cerradas), 1) if len(cerradas) else None
    )

    vencidas_abiertas = abiertas[abiertas["Semáforo"].str.startswith("Vencida", na=False)]

    # Efectividad por gerente
    filas_gerente = []
    for gerente, grupo in df.groupby("Gerente"):
        cerr = grupo[grupo["Estado Google Tasks"] == "Completada"]
        cerr_ok = cerr[cerr["Semáforo"] == "Cerrada en tiempo"]
        pct = round(100 * len(cerr_ok) / len(cerr), 1) if len(cerr) else None
        filas_gerente.append({
            "Gerente": gerente,
            "Total tareas": len(grupo),
            "Abiertas": len(grupo[grupo["Estado Google Tasks"] == "Pendiente"]),
            "Cerradas": len(cerr),
            "Cerradas en tiempo": len(cerr_ok),
            "% Efectividad": pct if pct is not None else "N/D",
            "Vencidas (abiertas)": len(grupo[
                (grupo["Estado Google Tasks"] == "Pendiente") &
                (grupo["Semáforo"].str.startswith("Vencida", na=False))
            ]),
        })
    df_gerente = pd.DataFrame(filas_gerente).sort_values("Gerente")

    # Distribución de semáforo
    conteo_semaforo = df["Semáforo"].value_counts().to_dict()

    # Distribución por categoría (Lista)
    conteo_categoria = df["Lista"].value_counts().to_dict()

    # Aging de tareas vencidas y abiertas (días de atraso)
    buckets_aging = {"0-3 días": 0, "4-7 días": 0, "8+ días": 0}
    for _, row in vencidas_abiertas.iterrows():
        fv = row["Fecha de vencimiento"]
        if fv is None:
            continue
        dias = (hoy - fv).days
        if dias <= 3:
            buckets_aging["0-3 días"] += 1
        elif dias <= 7:
            buckets_aging["4-7 días"] += 1
        else:
            buckets_aging["8+ días"] += 1

    return {
        "total_tareas": len(df),
        "total_gerentes": df["Gerente"].nunique(),
        "total_abiertas": len(abiertas),
        "total_cerradas": len(cerradas),
        "pct_efectividad_global": pct_efectividad_global,
        "total_vencidas": len(vencidas_abiertas),
        "df_gerente": df_gerente,
        "conteo_semaforo": conteo_semaforo,
        "conteo_categoria": conteo_categoria,
        "buckets_aging": buckets_aging,
    }


def _escribir_tabla(hoja, fila_inicio, col_inicio, headers, filas):
    """Escribe una tabla simple con encabezado estilizado, devuelve la
    fila siguiente a la última escrita."""
    for j, h in enumerate(headers):
        c = hoja.cell(row=fila_inicio, column=col_inicio + j, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10.5)
        c.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, fila in enumerate(filas, start=1):
        for j, valor in enumerate(fila):
            c = hoja.cell(row=fila_inicio + i, column=col_inicio + j, value=valor)
            c.font = Font(name="Arial", size=10)
    return fila_inicio + len(filas) + 1


def _tarjeta_kpi(hoja, fila, col, titulo, valor, color=NAVY):
    hoja.cell(row=fila, column=col, value=titulo).font = Font(name="Arial", size=10, color="6B7280")
    celda_valor = hoja.cell(row=fila + 1, column=col, value=valor)
    celda_valor.font = Font(name="Arial", size=20, bold=True, color=color)


def generar_excel(df: pd.DataFrame) -> bytes:
    metricas = calcular_metricas(df)
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # ---------- Hoja 1: Tareas (detalle) ----------------------------
        df.to_excel(writer, sheet_name="Tareas", index=False)
        hoja_tareas = writer.sheets["Tareas"]

        for col_idx, col_name in enumerate(df.columns, start=1):
            celda = hoja_tareas.cell(row=1, column=col_idx)
            celda.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            celda.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
            celda.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, col_name in enumerate(df.columns, start=1):
            letra = get_column_letter(col_idx)
            max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].fillna("")])
            hoja_tareas.column_dimensions[letra].width = min(max_len + 3, 45)
            for row_idx in range(2, len(df) + 2):
                hoja_tareas.cell(row=row_idx, column=col_idx).font = Font(name="Arial", size=10.5)

        hoja_tareas.freeze_panes = "A2"
        hoja_tareas.auto_filter.ref = hoja_tareas.dimensions
        hoja_tareas.sheet_view.showGridLines = False

        # ---------- Hoja 2: Dashboard -------------------------------------
        hoja_dash = writer.book.create_sheet("Dashboard", 0)  # la deja primera
        for letra, ancho in zip("ABCDE", [22, 16, 16, 16, 16]):
            hoja_dash.column_dimensions[letra].width = ancho
        hoja_dash.column_dimensions["F"].width = 17.17
        hoja_dash.column_dimensions["G"].width = 17.33
        hoja_dash.sheet_view.showGridLines = False

        hoja_dash["A1"] = "Dashboard Ejecutivo · Reporte de Tareas"
        hoja_dash["A1"].font = Font(name="Arial", size=16, bold=True, color=NAVY)
        hoja_dash["A2"] = f"Generado el {date.today().strftime('%d/%m/%Y')}"
        hoja_dash["A2"].font = Font(name="Arial", size=10, color="6B7280")

        # --- Tarjetas KPI (resumen general) ---
        _tarjeta_kpi(hoja_dash, 4, 1, "Total de tareas", metricas["total_tareas"])
        _tarjeta_kpi(hoja_dash, 4, 2, "Gerentes con tareas", metricas["total_gerentes"])
        _tarjeta_kpi(hoja_dash, 4, 3, "Abiertas", metricas["total_abiertas"])
        _tarjeta_kpi(hoja_dash, 4, 4, "Cerradas", metricas["total_cerradas"])
        _tarjeta_kpi(
            hoja_dash, 4, 5, "% Efectividad global",
            f"{metricas['pct_efectividad_global']}%" if metricas["pct_efectividad_global"] is not None else "N/D",
            color="1A8F3C",
        )
        _tarjeta_kpi(hoja_dash, 4, 6, "Vencidas (hoy)", metricas["total_vencidas"], color=RED)

        fila = 8

        # --- Tabla + gráfica: efectividad por gerente ---
        hoja_dash.cell(row=fila, column=1, value="Efectividad por gerente").font = Font(
            name="Arial", size=12, bold=True, color=NAVY
        )
        fila += 1
        df_g = metricas["df_gerente"]
        headers_g = list(df_g.columns)
        filas_g = df_g.values.tolist()
        fila_fin_tabla_g = _escribir_tabla(hoja_dash, fila, 1, headers_g, filas_g)

        fila = fila_fin_tabla_g + 2

        # --- Tabla + gráfica: distribución por semáforo ---
        hoja_dash.cell(row=fila, column=1, value="Distribución por semáforo").font = Font(
            name="Arial", size=12, bold=True, color=NAVY
        )
        fila += 1
        fila_inicio_sem = fila
        filas_sem = [[k, v] for k, v in metricas["conteo_semaforo"].items()]
        fila = _escribir_tabla(hoja_dash, fila, 1, ["Semáforo / Estado", "Tareas"], filas_sem)

        fila += 1

        # --- Tabla: distribución por categoría ---
        hoja_dash.cell(row=fila, column=1, value="Tareas por categoría").font = Font(
            name="Arial", size=12, bold=True, color=NAVY
        )
        fila += 1
        filas_cat = [[k, v] for k, v in sorted(metricas["conteo_categoria"].items())]
        fila = _escribir_tabla(hoja_dash, fila, 1, ["Categoría", "Tareas"], filas_cat)

        fila += 1

        # --- Tabla: antigüedad de tareas vencidas (aging) ---
        hoja_dash.cell(row=fila, column=1, value="Antigüedad de tareas vencidas (abiertas)").font = Font(
            name="Arial", size=12, bold=True, color=NAVY
        )
        fila += 1
        filas_aging = [[k, v] for k, v in metricas["buckets_aging"].items()]
        fila = _escribir_tabla(hoja_dash, fila, 1, ["Días de atraso", "Tareas"], filas_aging)

    return buffer.getvalue()


# --------------------------------------------------------------
# Evaluación por periodo (reglas del Director)
# --------------------------------------------------------------
def calcular_evaluacion(estado, fecha_vencimiento, fecha_terminacion, inicio, fin):
    """
    Tareas pendientes:
      - vencimiento <= fin del periodo -> "X" (debió cerrarse y no se cerró)
      - vencimiento >  fin del periodo -> "N.A." (aún no le tocaba)
    Tareas terminadas, se evalúan solo si:
      (a) su vencimiento cae dentro del periodo, o
      (b) se cerraron dentro del periodo Y su vencimiento era anterior al periodo
      Si se evalúan: terminación <= vencimiento -> "OK", si no -> "X" (se cerró tarde)
      Si no aplica ninguna condición -> "N.A."
    """
    if estado == "Pendiente":
        if fecha_vencimiento is None:
            return "N.A."
        return "X" if fecha_vencimiento <= fin else "N.A."
    elif estado == "Completada":
        if fecha_vencimiento is None:
            return "N.A."
        aplica_regla1 = inicio <= fecha_vencimiento <= fin
        aplica_regla2 = (
            fecha_terminacion is not None
            and inicio <= fecha_terminacion <= fin
            and fecha_vencimiento < inicio
        )
        if aplica_regla1 or aplica_regla2:
            if fecha_terminacion is None:
                return "N.A."
            return "OK" if fecha_terminacion <= fecha_vencimiento else "X"
        return "N.A."
    return "N.A."


def generar_excel_evaluacion(df: pd.DataFrame, inicio: date, fin: date) -> bytes:
    df_eval = df.copy()
    df_eval["Evaluación"] = df_eval.apply(
        lambda r: calcular_evaluacion(
            r["Estado Google Tasks"], r["Fecha de vencimiento"], r["Fecha de completado"], inicio, fin
        ),
        axis=1,
    )

    # Resumen de cumplimiento por gerente (excluye N.A. del % de cumplimiento)
    filas_resumen = []
    for gerente, grupo in df_eval.groupby("Gerente"):
        ok = (grupo["Evaluación"] == "OK").sum()
        x = (grupo["Evaluación"] == "X").sum()
        na = (grupo["Evaluación"] == "N.A.").sum()
        evaluadas = ok + x
        pct = round(100 * ok / evaluadas, 1) if evaluadas else "N/D"
        filas_resumen.append({
            "Gerente": gerente, "Evaluadas": evaluadas, "OK": ok, "X": x,
            "N.A.": na, "% Cumplimiento": pct,
        })
    df_resumen = pd.DataFrame(filas_resumen).sort_values("Gerente")

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        columnas_detalle = ["Gerente", "Lista", "Tarea", "Fecha de vencimiento",
                             "Estado Google Tasks", "Fecha de completado", "Evaluación"]
        df_eval[columnas_detalle].to_excel(writer, sheet_name="Evaluación", index=False, startrow=5)
        hoja = writer.sheets["Evaluación"]
        hoja.sheet_view.showGridLines = False

        hoja["A1"] = "Evaluación de Tareas por Gerente"
        hoja["A1"].font = Font(name="Arial", size=16, bold=True, color=NAVY)
        hoja["A2"] = f"Periodo evaluado: {inicio.strftime('%d/%m/%Y')} — {fin.strftime('%d/%m/%Y')}"
        hoja["A2"].font = Font(name="Arial", size=11, color="6B7280")
        hoja["A3"] = "Nota: la fecha de elaboración no está disponible (Google Tasks no la expone)."
        hoja["A3"].font = Font(name="Arial", size=9, italic=True, color="9CA3AF")

        for col_idx, col_name in enumerate(columnas_detalle, start=1):
            celda = hoja.cell(row=6, column=col_idx)
            celda.font = Font(name="Arial", bold=True, color="FFFFFF", size=10.5)
            celda.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
            celda.alignment = Alignment(horizontal="center", vertical="center")
        for col_idx, col_name in enumerate(columnas_detalle, start=1):
            letra = get_column_letter(col_idx)
            valores = df_eval[col_name].fillna("")
            max_len = max([len(str(col_name))] + [len(str(v)) for v in valores])
            hoja.column_dimensions[letra].width = min(max_len + 3, 40)
            for row_idx in range(7, len(df_eval) + 7):
                hoja.cell(row=row_idx, column=col_idx).font = Font(name="Arial", size=10)
        hoja.freeze_panes = "A7"
        hoja.auto_filter.ref = f"A6:{get_column_letter(len(columnas_detalle))}{6 + len(df_eval)}"

        # ---- Resumen de cumplimiento por gerente, a la derecha ----
        col_resumen = len(columnas_detalle) + 2
        hoja.cell(row=6, column=col_resumen, value="Resumen de cumplimiento por gerente").font = Font(
            name="Arial", size=12, bold=True, color=NAVY
        )
        headers_r = list(df_resumen.columns)
        filas_r = df_resumen.values.tolist()
        _escribir_tabla(hoja, 7, col_resumen, headers_r, filas_r)
        for j, ancho in enumerate([20, 12, 10, 10, 10, 16]):
            hoja.column_dimensions[get_column_letter(col_resumen + j)].width = ancho

    return buffer.getvalue()


# --------------------------------------------------------------
# UI
# --------------------------------------------------------------
if _mostrar_debug:
    mostrar_debug_credenciales()

LISTAS_PERMITIDAS_DEFAULT = ["AA", "AB", "BA", "BB"]

with st.container(border=True):
    st.markdown("**📋 Categorías a incluir**")
    listas_seleccionadas = st.multiselect(
        "Solo se consolidan las tareas capturadas en estas listas:",
        options=LISTAS_PERMITIDAS_DEFAULT,
        default=LISTAS_PERMITIDAS_DEFAULT,
        label_visibility="collapsed",
        help="Las tareas en cualquier otra lista (ej. 'Mis tareas') se excluyen del reporte.",
    )

st.write("")
generar = st.button("🔄  Generar reporte", use_container_width=True)

if generar:
    n_gerentes = len(st.secrets.get("gerentes", []))
    with st.spinner(f"Leyendo Google Tasks de {n_gerentes} cuenta(s)..."):
        try:
            df, errores = extraer_todas_las_cuentas()
        except Exception as e:
            st.error(f"No se pudo generar el reporte: {e}")
            st.stop()

    if errores:
        for err in errores:
            st.warning(f"⚠️ No se pudo leer la cuenta de {err}")

    # Filtrar solo las listas seleccionadas (AA, AB, BA, BB por default)
    if not df.empty and listas_seleccionadas:
        total_antes = len(df)
        df = df[df["Lista"].isin(listas_seleccionadas)]
        excluidas = total_antes - len(df)
        if excluidas:
            st.caption(f"({excluidas} tarea(s) excluida(s) por no estar en las categorías seleccionadas)")

    if df.empty:
        st.warning("No se encontraron tareas en ninguna cuenta con las categorías seleccionadas.")
    else:
        st.success(f"✅ Se encontraron **{len(df)} tareas** de **{df['Gerente'].nunique()} gerente(s)**.")
        st.dataframe(df, use_container_width=True, hide_index=True)

        excel_bytes = generar_excel(df)
        st.download_button(
            "⬇️ Descargar Excel",
            data=excel_bytes,
            file_name=f"reporte_tareas_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

st.write("")
st.markdown("---")

with st.container(border=True):
    st.markdown("**🗓️ Evaluación de tareas por periodo**")
    st.caption("Genera un reporte independiente con la evaluación de cumplimiento de cada gerente, "
               "según las reglas del Director, para el rango de fechas que elijas.")
    col1, col2 = st.columns(2)
    fecha_inicio = col1.date_input("Fecha de inicio", value=date.today().replace(day=1))
    fecha_termino = col2.date_input("Fecha de término", value=date.today())

    generar_eval = st.button("📊  Generar evaluación", use_container_width=True)

    if generar_eval:
        if fecha_inicio > fecha_termino:
            st.error("La fecha de inicio no puede ser posterior a la fecha de término.")
        else:
            n_gerentes = len(st.secrets.get("gerentes", []))
            with st.spinner(f"Leyendo Google Tasks de {n_gerentes} cuenta(s)..."):
                try:
                    df_eval_base, errores_eval = extraer_todas_las_cuentas()
                except Exception as e:
                    st.error(f"No se pudo generar la evaluación: {e}")
                    st.stop()

            if errores_eval:
                for err in errores_eval:
                    st.warning(f"⚠️ No se pudo leer la cuenta de {err}")

            if not df_eval_base.empty and listas_seleccionadas:
                df_eval_base = df_eval_base[df_eval_base["Lista"].isin(listas_seleccionadas)]

            if df_eval_base.empty:
                st.warning("No hay tareas para evaluar con las categorías seleccionadas.")
            else:
                excel_eval_bytes = generar_excel_evaluacion(df_eval_base, fecha_inicio, fecha_termino)
                st.success(f"✅ Evaluación generada para el periodo "
                           f"{fecha_inicio.strftime('%d/%m/%Y')} – {fecha_termino.strftime('%d/%m/%Y')}.")
                st.download_button(
                    "⬇️ Descargar evaluación",
                    data=excel_eval_bytes,
                    file_name=f"evaluacion_{fecha_inicio.isoformat()}_a_{fecha_termino.isoformat()}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
